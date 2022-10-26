import json
import pandas as pd
import numpy as np
import openpyxl
import sys, os
import glob
import pyexcel
import win32com.client
from slugify import slugify # pip install python-slugify
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 

# config
root_dir = r'D:\Projects\lct_8\Исходные данные\Исходные данные\Сметы для конкурса\Сметы для задачи исходные'
all_files = []


# Функция для рекурсивного обхода дир и субдир для поиска xls* файлоа
def enumerate_files(root_dir):
    all_files = []
    for filename in glob.iglob(root_dir + '**/**', recursive=True):
        if 'xls' in filename and '~$' not in filename:
            all_files.append(filename)

    return all_files

# convert xls to xlsx
def conv_xls_xlsx(infile):
    dot = infile.find('.')
    end = infile[dot:]
    outfile = infile[0:dot] + ".xlsx"
    pyexcel.save_as(file_name=infile, dest_file_name=outfile)
    return outfile
    
# Основная функция для извлечения информации из xls файлов в json файлы
def extract_data(all_files):    
    json_list = []
    
    for xls_file in all_files:
        all_items = []
        wb = None
        
        try:
            wb = openpyxl.load_workbook(xls_file, data_only=True) # открываем xls файл в режиме даты
        except:
            print(xls_file, ' более старой версии и не доддерживается openpyxl.')
            newfile = conv_xls_xlsx(xls_file)
            wb = openpyxl.load_workbook(newfile, data_only=True) # открываем xls файл в режиме даты
            #continue
        
        filename_s = os.path.basename(xls_file)  # получаем имя файла из пути   
        filename_slug = slugify(filename_s, separator='_')+'.json'  # получаем безопасное имя файла  
        print('Текущий файл:', filename_s, filename_slug)

        #sheets_vedomost = ['Дефектная ведомость', 'Ведомость объемов работ']
        sheets_vedomost = wb.get_sheet_names() # получаем все вкладки в документе

        # записываем имя файла
        all_items.append(filename_s)

        # пробегаемся по всем вкладкам файла
        for idx, sh in enumerate(sheets_vedomost):
            #if idx>10: break
            
            try:
                
                ################################################################## 
                # для каждой вкладки ведомостей
                if 'ведомость' in sh.lower():          
                        all_items = []      
                        sheet = wb.get_sheet_by_name(sh)
                        sheet_size = sheet.max_row
                        c_max = 'A'+str(sheet_size)
                        
                        # записываем имя вкладки
                        all_items.append(sh) 
                        # для каждой строчки
                        for rowOfCellObjects in sheet['A1':c_max]:
                            # для каждого элемента
                            for cellObj in rowOfCellObjects:
                                if 'Локальная смета:' in str(cellObj.value) or 'Раздел:' in str(cellObj.value) or 'Подраздел:' in str(cellObj.value):
                                    all_items.append(cellObj.value) 
                                # получаем id
                                if cellObj.value != None and len(str(cellObj.value)) < 4:
                                    # print(cellObj.coordinate, cellObj.value)
                                    # достаём дельту с координат
                                    d = str(cellObj.coordinate)[1:]
                                    # если нет названия работ то пропускаем
                                    if len(str(sheet['B'+d].value)) < 10: continue
                                    
                                    main_work_dict = {}
                                    
                                    # достаём значения
                                    id = sheet['A'+d].value
                                    work_name = sheet['B'+d].value
                                    unit = sheet['C'+d].value
                                    amount = sheet['D'+d].value    
                                                        
                                    # определяем суб работу по наличии запятой
                                    if ',' in str(sheet['A'+d].value):
                                            sub_work_dict = {}
                                            sub_work_dict["Number"] = id  
                                            sub_work_dict["Name"] = work_name  
                                            sub_work_dict["Unit"] = unit  
                                            sub_work_dict["Amount"] = amount  
                                            main_work_dict["Sub_works"] = sub_work_dict
                                    else:                     
                                            main_work_dict["Number"] = id  
                                            main_work_dict["Name"] = work_name  
                                            main_work_dict["Unit"] = unit  
                                            main_work_dict["Amount"] = amount  
                                            main_work_dict["Sub_works"] = []
                                            main_work_dict["Related_works"] = []
                                            
                                    all_items.append(main_work_dict)
                                    #print(all_items)
                        
                        # дампим в файл
                        with open('data/'+filename_slug, "w") as f:
                            json.dump(all_items, f)            
                    
                
                ##################################################################  
                # для каждой вкладки ресурсов      
                if 'ресурсов' in sh.lower():    

                    all_items = []
                    # Расчет стоимости ресурсов
                    sheet = wb.get_sheet_by_name(sh)
                    sheet_size = sheet.max_row
                    c_max = 'A'+str(sheet_size)

                    # записываем имя вкладки
                    all_items.append(sh) 
                    # для каждой строчки
                    for rowOfCellObjects in sheet['A1':c_max]:
                        # для кадого элемента
                        for cellObj in rowOfCellObjects:
                            # получаем id
                            if cellObj.value != None and \
                                '-' in str(cellObj.value) and '.' in str(cellObj.value) or \
                                'коммерческое предложение' in str(cellObj.value):
                                # print(cellObj.coordinate, cellObj.value)
                                # достаём дельту с координат
                                d = str(cellObj.coordinate)[1:]
                                # если нет названия работ то пропускаем
                                if len(str(sheet['B'+d].value)) < 10: continue
                                
                                main_work_dict = {}
                                
                                # достаём значения
                                number = sheet['A'+d].value
                                work_name = sheet['B'+d].value
                                unit = sheet['C'+d].value
                                amount = sheet['D'+d].value    
                                price = sheet['E'+d].value    
                                curr_p = sheet['F'+d].value    
                                                    
                                # определяем суб работу по наличии запятой
                                main_work_dict["Number"] = number  
                                main_work_dict["Name"] = work_name  
                                main_work_dict["Unit"] = unit  
                                main_work_dict["Price"] = price  
                                main_work_dict["Cprice"] = curr_p  
                                        
                                all_items.append(main_work_dict)
                                #print(all_items)
                                
                    # дампим в файл
                    with open('data/'+filename_slug, "w") as f:
                        json.dump(all_items, f) 
                
                json_list.append(filename_slug)  
            except:
                    print(sh, ' возникла ошибка при обработке... в файле', filename_s, filename_slug)  
    return json_list                


# Получаем список xls* файлов
all_files = enumerate_files(root_dir)
result_list = extract_data(all_files)
total = len(result_list)
print('Всего обработанно: ', total)
